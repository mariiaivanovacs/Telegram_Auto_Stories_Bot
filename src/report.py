import io
import logging

logger = logging.getLogger(__name__)

# Try to import openpyxl at module level
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError as e:
    logger.warning("openpyxl not available for Excel export: %s", e)
    HAS_OPENPYXL = False


def build_price_list(price_results: list[dict], template: str) -> str:
    """
    Fill the price list template placeholders with current prices.
    Missing prices render as '—'.
    """
    values: dict[str, str] = {}
    for r in price_results:
        price = r.get("calculated_price")
        values[r["template_key"]] = _fmt(price) if price is not None else "—"

    try:
        return template.format(**values)
    except KeyError as e:
        logger.warning("Template placeholder %s not found in pricing results", e)
        return template


def build_report(
    price_results: list[dict],
    unavailable_channels: list[str],
    all_channel_names: list[str],
    started_at: str,
    errors: list[str],
) -> str:
    """
    Build the run report Telegram message.

    Args:
        price_results:       output of pricing.calculate_prices()
        unavailable_channels: channel usernames that could not be fetched
        all_channel_names:   all configured channel usernames
        started_at:          ISO-8601 timestamp of run start
        errors:              list of error strings collected during the run
    """
    ts = started_at[:16].replace("T", " ")

    found = [r for r in price_results if not r.get("price_kept")]
    missing = [r for r in price_results if r.get("price_kept")]

    lines: list[str] = [f"📊 Отчёт запуска — {ts}", ""]

    lines.append(f"✅ Найдено цен: {len(found)} / {len(price_results)}")
    if missing:
        lines.append("❌ Отсутствуют: " + ", ".join(r["canonical_name"] for r in missing))

    lines.append("")
    for ch in all_channel_names:
        status = "НЕДОСТУПЕН ⚠️" if ch in unavailable_channels else "ОК"
        lines.append(f"Канал: {ch} — {status}")

    if found:
        lines.append("")
        lines.append("Изменения цен:")
        for r in found:
            old = r.get("old_price")
            new = r.get("calculated_price")
            delta = r.get("price_delta", 0)

            if old is None:
                line = f"• {r['canonical_name']}: {_fmt(new)} RUB (первый запуск)"
            elif delta == 0:
                line = f"• {r['canonical_name']}: {_fmt(new)} RUB (без изменений)"
            else:
                sign = "−" if delta < 0 else "+"
                line = f"• {r['canonical_name']}: {_fmt(old)} → {_fmt(new)} ({sign}{_fmt(abs(delta))})"

            if r.get("is_large_change"):
                line += " ⚠️ БОЛЬШОЕ ИЗМЕНЕНИЕ"
            lines.append(line)

    lines.append("")
    if errors:
        lines.append("Ошибки:")
        for e in errors:
            lines.append(f"  • {e}")
    else:
        lines.append("Ошибки: нет")

    return "\n".join(lines)


def build_competition_report_excel(
    run: dict,
    rows: list[dict],
    history: list[dict] | None = None,
) -> bytes:
    """
    Generate an Excel (.xlsx) competition price report with 3 sheets:
      1. Сводка        — run summary
      2. Текущий запуск — current run with full colour-coded delta highlighting
      3. История 30 дней — flat history table, colour-coded the same way
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    # Colour palette (RGB hex without leading #)
    _HEADER_BG   = "1F4E79"
    _RED_DARK    = "FF6666"   # increase > 3 000 ₽
    _RED_LIGHT   = "FFD9CC"   # increase 0–3 000 ₽
    _GREEN_DARK  = "66CC88"   # drop > 3 000 ₽
    _GREEN_LIGHT = "CCFFDD"   # drop 0–3 000 ₽
    _GREY        = "F2F2F2"   # no data / price kept

    def _delta_fill(delta):
        if delta is None:
            return PatternFill(start_color=_GREY, end_color=_GREY, fill_type="solid")
        if delta > 3_000:
            return PatternFill(start_color=_RED_DARK,   end_color=_RED_DARK,   fill_type="solid")
        if delta > 0:
            return PatternFill(start_color=_RED_LIGHT,  end_color=_RED_LIGHT,  fill_type="solid")
        if delta < -3_000:
            return PatternFill(start_color=_GREEN_DARK,  end_color=_GREEN_DARK,  fill_type="solid")
        if delta < 0:
            return PatternFill(start_color=_GREEN_LIGHT, end_color=_GREEN_LIGHT, fill_type="solid")
        return None  # delta == 0: no fill

    def _apply_header(ws, header_row: list[str]) -> None:
        ws.append(header_row)
        bg  = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
        fnt = Font(bold=True, color="FFFFFF")
        for cell in ws[ws.max_row]:
            cell.fill = bg
            cell.font = fnt
            cell.alignment = Alignment(horizontal="center")

    def _autofit(ws) -> None:
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w, 40)

    def _delta_str(delta) -> str:
        if delta is None:
            return "—"
        sign = "+" if delta > 0 else ""
        return f"{sign}{delta:,}".replace(",", " ")

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Сводка"
    run_date = (run.get("started_at") or "")[:16].replace("T", " ")
    total = (run.get("products_found") or 0) + (run.get("products_missing") or 0)
    for label, value in [
        ("Дата запуска", run_date),
        ("Найдено цен",  f"{run.get('products_found', 0)} / {total}"),
        ("Пропущено",    run.get("products_missing", 0)),
        ("", ""),
        ("Обозначение",   "Цвет"),
        ("Рост > 3 000 ₽",    "Красный"),
        ("Рост 1–3 000 ₽",    "Светло-красный"),
        ("Снижение 1–3 000 ₽","Светло-зелёный"),
        ("Снижение > 3 000 ₽", "Зелёный"),
        ("Нет данных",         "Серый"),
    ]:
        ws1.append([label, str(value)])
        ws1.cell(row=ws1.max_row, column=1).font = Font(bold=True)
    _autofit(ws1)

    # ── Sheet 2: Current run ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Текущий запуск")
    _apply_header(ws2, ["Товар", "Категория", "Наша цена (₽)", "Цена конкур. (₽)",
                         "Канал", "Изменение (₽)", "Статус"])

    for r in rows:
        delta = r.get("price_delta")
        if r.get("price_kept"):
            status = "Нет данных"
        elif r.get("is_large_change"):
            status = "⚠️ Крупное изменение"
        elif delta and delta != 0:
            status = "Изменение"
        else:
            status = "✅ Без изменений"

        ws2.append([
            r.get("display_name", ""),
            r.get("category", ""),
            r.get("calculated_price") or "—",
            r.get("competitor_price") or "—",
            f"@{r['source_channel']}" if r.get("source_channel") else "—",
            _delta_str(delta),
            status,
        ])
        fill = _delta_fill(delta if not r.get("price_kept") else None)
        if fill:
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
    _autofit(ws2)

    # ── Sheet 3: 30-day history ────────────────────────────────────────────────
    ws3 = wb.create_sheet("История 30 дней")
    _apply_header(ws3, ["Дата", "Товар", "Категория", "Цена конкур. (₽)",
                         "Наша цена (₽)", "Изменение (₽)", "Изм. %", "Канал", "Статус"])

    for r in (history or []):
        delta     = r.get("price_delta")
        our_price = r.get("calculated_price")
        comp      = r.get("competitor_price")
        kept      = r.get("price_kept")

        pct = ""
        if delta and our_price and our_price - delta:
            pct = f"{delta / (our_price - delta) * 100:+.1f}%"

        if kept:
            status = "Нет данных"
        elif r.get("is_large_change"):
            status = "⚠️ Крупное"
        elif delta and delta != 0:
            status = "Изменение"
        else:
            status = "Без изменений"

        ws3.append([
            (r.get("started_at") or "")[:16].replace("T", " "),
            r.get("display_name", ""),
            r.get("category", ""),
            comp or "—",
            our_price or "—",
            _delta_str(delta),
            pct,
            f"@{r['source_channel']}" if r.get("source_channel") else "—",
            status,
        ])
        fill = _delta_fill(delta if not kept else None)
        if fill:
            for cell in ws3[ws3.max_row]:
                cell.fill = fill
    _autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _fmt(n: int | None) -> str:
    if n is None:
        return "—"
    return f"{n:,}".replace(",", " ")
